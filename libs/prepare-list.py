# -*- coding: utf-8 -*-
"""Basic implementation of command line tool that receives a list of unnamed arguments"""

import os
import re
import json
import shutil
from typing import List

import typer
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - runtime import
    load_workbook = None

try:
    from sqlalchemy import create_engine
except Exception:
    create_engine = None


# Instantiate the typer library
app = typer.Typer()


# define the function for the command line
@app.command()
def main(
    input_folder: str = typer.Option(
        None,
        "-i",
        "--input",
        help="Input folder containing Excel files (env: PREP_INPUT)",
    ),
    output_folder: str = typer.Option(
        None,
        "-o",
        "--output",
        help="Output folder where processed files are written (env: PREP_OUTPUT)",
    ),
    schema_file: str = typer.Option(
        None, "-s", "--schema", help="Path to schema.json (env: PREP_SCHEMA)"
    ),
    sql_file: str = typer.Option(
        None,
        "-q",
        "--sql",
        help="Path to SQL file to execute (env: PREP_SQL)",
    ),
    db_uri: str = typer.Option(
        None,
        "-d",
        "--db-uri",
        help="Database URI (SQLAlchemy style) used to run the SQL and fetch rows (env: DATABASE_URL)",
    ),
) -> None:
    """Copy Excel files from `input_folder` to `output_folder`, then populate them
    using `schema.json` and rows returned from `sql_file` executed against `db_uri`.

    The SQL is expected to return columns including: `partid`, `sku`, `description`,
    `price`, `effective_date`, and `category` (category like 'RTL-1' or 'RTL-1' inside
    brackets depending on your SQL). The function will pivot categories into columns
    and map fields to the Excel columns according to `schema.json`.
    """
    # load environment variables from .env (if present)
    load_dotenv()

    # resolve defaults from environment when CLI args not provided
    default_input = os.getenv("PREP_INPUT", os.path.join("lists", "templates"))
    default_output = os.getenv("PREP_OUTPUT", os.path.join("lists", "output"))
    default_schema = os.getenv("PREP_SCHEMA", "lists/schema.json")
    default_sql = os.getenv("PREP_SQL", "sql-scripts/extract_query.sql")
    default_db = os.getenv("DATABASE_URL")

    input_folder = input_folder or default_input
    output_folder = output_folder or default_output
    schema_file = schema_file or default_schema
    sql_file = sql_file or default_sql
    db_uri = db_uri or default_db

    # Load lookup CSV (first CSV found in PREP_LOOKUP) into a dict keyed by Part #
    lookup_dir = os.getenv("PREP_LOOKUP", os.path.join("lists", "lookup"))
    lookup_map = {}
    try:
        p = Path(lookup_dir)
        lookup_file = None
        if p.exists() and p.is_dir():
            for f in sorted(p.glob("*.csv")):
                lookup_file = f
                break

        if lookup_file:
            import csv

            with lookup_file.open(newline="", encoding="utf-8") as fh:
                reader = csv.DictReader(fh)

                # prefer explicit "Part #" column as the key; fall back to common alternatives
                key_col = "Part #"

                for row in reader:
                    print(row)
                    raw_key = row.get(key_col)
                    if not raw_key:
                        continue
                    key_norm = str(raw_key).strip()

                    partid_val = row.get("\ufeffPart ID")
                    desc_val = row.get("Description")
                    entry = {
                        "partid": str(partid_val).strip(),
                        "description": str(desc_val).strip(),
                    }
                    # simple direct mapping: key is the raw Part # value (stripped)
                    lookup_map[key_norm] = entry

            typer.echo(f"Loaded lookup file: {lookup_file} ({len(lookup_map)} entries)")
        else:
            typer.echo(
                f"No lookup CSV found in {lookup_dir}; continuing without lookup"
            )
    except Exception as exc:
        typer.echo(f"Failed to load lookup CSV: {exc}")

    in_root = Path(input_folder)
    out_root = Path(output_folder)

    if not in_root.exists() or not in_root.is_dir():
        typer.echo(f"Input folder does not exist or is not a directory: {in_root}")
        raise typer.Exit(2)

    out_root.mkdir(parents=True, exist_ok=True)

    # load schema
    schema_path = Path(schema_file)
    if not schema_path.exists():
        typer.echo(f"Schema file not found: {schema_path}")
        raise typer.Exit(3)

    with schema_path.open("r", encoding="utf-8") as fh:
        schemas = json.load(fh)

    # read SQL
    sql_path = Path(sql_file)
    if not sql_path.exists():
        typer.echo(f"SQL file not found: {sql_path}")
        raise typer.Exit(4)

    sql_text = sql_path.read_text()

    if not db_uri:
        typer.echo("Database URI (--db-uri) is required to execute the SQL.")
        raise typer.Exit(5)

    if create_engine is None:
        typer.echo(
            "sqlalchemy is required to run queries. Install it in your environment."
        )
        raise typer.Exit(6)

    engine = create_engine(db_uri)

    # execute SQL and load dataframe
    try:
        df_sql = pd.read_sql(sql_text, con=engine)
    except Exception as exc:
        typer.echo(f"Failed to execute SQL: {exc}")
        raise typer.Exit(7)

    # normalize SQL dataframe column names to lowercase for consistent lookups
    df_sql.columns = [str(c).lower() for c in df_sql.columns]

    # process each excel file under input folder
    excel_suffixes = {".xls", ".xlsx"}
    for src in sorted(in_root.rglob("*")):
        if not src.is_file() or src.suffix.lower() not in excel_suffixes:
            continue

        rel = src.relative_to(in_root)
        rel_str = str(rel).replace("\\", "/")
        print(f"Processing file: {rel_str}")

        # find corresponding schema entry (schemas is a list of dicts)
        match = None
        for entry in schemas:
            if entry.get("filename") == rel_str:
                match = entry
                break

        if match is None:
            typer.echo(f"No schema entry for: {rel_str} - skipping")
            continue

        # copy file to output preserving relative path
        dest = out_root / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dest)

        # open workbook
        if load_workbook is None:
            typer.echo(
                "openpyxl is required to write Excel files. Install it in your environment."
            )
            raise typer.Exit(8)

        wb = load_workbook(dest)
        ws = wb.active

        starting_row = int(match.get("starting_row", 1))

        # determine header row by looking in rows 1..starting_row (pick best match)
        schema_columns = match.get("columns", [])
        header_row = None
        best_score = -1
        for r in range(1, starting_row + 1):
            row_vals = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in ws[r]
            ]
            score = 0
            for col in schema_columns:
                name = col.get("columnname", "").strip()
                if name in row_vals:
                    score += 1
            if score > best_score:
                best_score = score
                header_row = r

        # build column name -> Excel column index mapping
        col_map = {}
        if header_row is not None and best_score > 0:
            row_vals = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in ws[header_row]
            ]
            for idx, val in enumerate(row_vals, start=1):
                for col in schema_columns:
                    if col.get("columnname", "").strip() == val:
                        col_map[col.get("columnname")] = idx
        else:
            # fallback: map schema columns in order starting at column 1
            for idx, col in enumerate(schema_columns, start=1):
                col_map[col.get("columnname")] = idx

        # detect partid and description columns in the Excel schema
        key_idx = 1
        partid_col_idx = 2
        desc_col_idx = 3

        # prepare rows to write from SQL results (no pivot required)
        # reset index to ensure iteration
        df_rows = df_sql.copy()

        # helper to evaluate dfcolumn expressions like '[DST-1]/0.9'
        expr_re = re.compile(r"\[([^\]]+)\](?:\s*([*/+-])\s*([0-9.]+))?")

        def eval_expr(expr: str, row: pd.Series):
            # direct column name (case-insensitive)
            if expr in row.index:
                return row.get(expr)
            if expr.lower() in row.index:
                return row.get(expr.lower())

            # expression with bracketed category
            m = expr_re.search(expr)
            if not m:
                return None
            cat = m.group(1)
            op = m.group(2)
            num = m.group(3)
            # try several casings for the category column
            val = None
            if cat in row.index:
                val = row.get(cat)
            elif cat.lower() in row.index:
                val = row.get(cat.lower())
            elif cat.upper() in row.index:
                val = row.get(cat.upper())
            try:
                if pd.isna(val):
                    return None
            except Exception:
                pass
            try:
                if op and num:
                    numf = float(num)
                    if op == "/":
                        return float(val) / numf
                    if op == "*":
                        return float(val) * numf
                    if op == "+":
                        return float(val) + numf
                    if op == "-":
                        return float(val) - numf
                return val
            except Exception:
                return val

        # start writing data rows beginning at starting_row
        write_row = starting_row
        for _, prow in df_rows.iterrows():
            for col in schema_columns:
                cname = col.get("columnname")
                dfield = col.get("dfcolumnname", "")

                # default empty cell if dfcolumnname is missing/null
                if not dfield:
                    val = None
                else:
                    dfield_l = dfield.strip()

                    def get_value(expr: str, row: pd.Series):
                        """Return a value for a simple expr which may be a column name or a [bracketed] category."""
                        if not expr:
                            return None
                        # direct lookup (case-insensitive)
                        if expr in row.index:
                            return row.get(expr)
                        if expr.lower() in row.index:
                            return row.get(expr.lower())
                        # bracketed category
                        m = expr_re.search(expr)
                        if m:
                            cat = m.group(1)
                            if cat in row.index:
                                return row.get(cat)
                            if cat.lower() in row.index:
                                return row.get(cat.lower())
                            if cat.upper() in row.index:
                                return row.get(cat.upper())
                        # literal number?
                        try:
                            return float(expr)
                        except Exception:
                            return None

                    # if contains '/', treat as a simple formula left/right
                    if "*" in dfield_l:
                        left, right = dfield_l.split("*", 1)
                        lval = get_value(left.strip(), prow)
                        rval = None
                        try:
                            rval = float(right.strip())
                        except Exception:
                            rval = get_value(right.strip(), prow)

                        try:
                            if lval is None or rval in (None, 0):
                                val = None
                            else:
                                val = float(lval) * float(rval)
                        except Exception:
                            val = None
                    else:
                        # simple lookup or bracket expression
                        val = get_value(dfield_l, prow)

                col_idx = col_map.get(cname)
                if col_idx is None:
                    # append to the end
                    col_idx = ws.max_column + 1
                    col_map[cname] = col_idx

                ws.cell(row=write_row, column=col_idx, value=val)

            # After writing all columns for this row, try to populate Part # and Description from lookup_map
            try:
                # determine key from the Excel partid column if present, otherwise from SQL row 'partid'
                key_val = None
                if key_idx is not None:
                    cell_val = ws.cell(row=write_row, column=key_idx).value
                    if cell_val is not None:
                        key_val = str(cell_val).strip()
                if not key_val:
                    # try SQL row field (lowercase normalized earlier)
                    if "partid" in df_rows.columns:
                        key_val = str(prow.get("partid") or "").strip()

                if key_val:
                    lookup_key = key_val.strip()
                    lm = lookup_map.get(lookup_key)

                    if lm:
                        if partid_col_idx is not None:
                            ws.cell(
                                row=write_row,
                                column=partid_col_idx,
                                value=lm.get("partid", key_val),
                            )
                        if desc_col_idx is not None:
                            ws.cell(
                                row=write_row,
                                column=desc_col_idx,
                                value=lm.get("description", ""),
                            )
            except Exception:
                # non-fatal: continue processing
                print("lookup failed for key:", key_val)
                pass

            write_row += 1

        wb.save(dest)
        typer.echo(f"Processed and wrote file: {dest}")


# ----------------------------------------------------------------
# Main
# ----------------------------------------------------------------
if __name__ == "__main__":
    app()
