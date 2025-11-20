# -*- coding: utf-8 -*-
"""Basic implementation of command line tool that receives a list of unnamed arguments"""

import os
import re
import json
import shutil
from typing import List

import typer
import pandas as pd
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - runtime import
    load_workbook = None

try:
    from sqlalchemy import create_engine
except Exception:
    create_engine = None


# Instantiate the typer library
app = typer.Typer()


# define the function for the command line
@app.command()
def main(
    input_folder: str = typer.Option(
        ..., "-i", "--input", help="Input folder containing Excel files"
    ),
    output_folder: str = typer.Option(
        ..., "-o", "--output", help="Output folder where processed files are written"
    ),
    schema_file: str = typer.Option(
        "prepare-lists/schema.json", "-s", "--schema", help="Path to schema.json"
    ),
    sql_file: str = typer.Option(
        "sql-scripts/extract_query.sql",
        "-q",
        "--sql",
        help="Path to SQL file to execute",
    ),
    db_uri: str = typer.Option(
        None,
        "-d",
        "--db-uri",
        help="Database URI (SQLAlchemy style) used to run the SQL and fetch rows",
    ),
) -> None:
    """Copy Excel files from `input_folder` to `output_folder`, then populate them
    using `schema.json` and rows returned from `sql_file` executed against `db_uri`.

    The SQL is expected to return columns including: `partid`, `sku`, `description`,
    `price`, `effective_date`, and `category` (category like 'RTL-1' or 'RTL-1' inside
    brackets depending on your SQL). The function will pivot categories into columns
    and map fields to the Excel columns according to `schema.json`.
    """
    in_root = Path(input_folder)
    out_root = Path(output_folder)

    if not in_root.exists() or not in_root.is_dir():
        typer.echo(f"Input folder does not exist or is not a directory: {in_root}")
        raise typer.Exit(2)

    out_root.mkdir(parents=True, exist_ok=True)

    # load schema
    schema_path = Path(schema_file)
    if not schema_path.exists():
        typer.echo(f"Schema file not found: {schema_path}")
        raise typer.Exit(3)

    with schema_path.open("r", encoding="utf-8") as fh:
        schemas = json.load(fh)

    # read SQL
    sql_path = Path(sql_file)
    if not sql_path.exists():
        typer.echo(f"SQL file not found: {sql_path}")
        raise typer.Exit(4)

    sql_text = sql_path.read_text()

    if not db_uri:
        typer.echo("Database URI (--db-uri) is required to execute the SQL.")
        raise typer.Exit(5)

    if create_engine is None:
        typer.echo(
            "sqlalchemy is required to run queries. Install it in your environment."
        )
        raise typer.Exit(6)

    engine = create_engine(db_uri)

    # execute SQL and load dataframe
    try:
        df_sql = pd.read_sql(sql_text, con=engine)
    except Exception as exc:
        typer.echo(f"Failed to execute SQL: {exc}")
        raise typer.Exit(7)

    # normalize SQL dataframe column names to lowercase for consistent lookups
    df_sql.columns = [str(c).lower() for c in df_sql.columns]

    # process each excel file under input folder
    excel_suffixes = {".xls", ".xlsx"}
    for src in sorted(in_root.rglob("*")):
        if not src.is_file() or src.suffix.lower() not in excel_suffixes:
            continue

        rel = src.relative_to(in_root)
        rel_str = str(rel).replace("\\", "/")
        print(f"Processing file: {rel_str}")

        # find corresponding schema entry (schemas is a list of dicts)
        match = None
        for entry in schemas:
            if entry.get("filename") == rel_str:
                match = entry
                break

        if match is None:
            typer.echo(f"No schema entry for: {rel_str} - skipping")
            continue

        # copy file to output preserving relative path
        dest = out_root / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dest)

        # open workbook
        if load_workbook is None:
            typer.echo(
                "openpyxl is required to write Excel files. Install it in your environment."
            )
            raise typer.Exit(8)

        wb = load_workbook(dest)
        ws = wb.active

        starting_row = int(match.get("starting_row", 1))

        # determine header row by looking in rows 1..starting_row (pick best match)
        schema_columns = match.get("columns", [])
        header_row = None
        best_score = -1
        for r in range(1, starting_row + 1):
            row_vals = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in ws[r]
            ]
            score = 0
            for col in schema_columns:
                name = col.get("columnname", "").strip()
                if name in row_vals:
                    score += 1
            if score > best_score:
                best_score = score
                header_row = r

        # build column name -> Excel column index mapping
        col_map = {}
        if header_row is not None and best_score > 0:
            row_vals = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in ws[header_row]
            ]
            for idx, val in enumerate(row_vals, start=1):
                for col in schema_columns:
                    if col.get("columnname", "").strip() == val:
                        col_map[col.get("columnname")] = idx
        else:
            # fallback: map schema columns in order starting at column 1
            for idx, col in enumerate(schema_columns, start=1):
                col_map[col.get("columnname")] = idx

        # prepare rows to write from SQL results (no pivot required)
        # reset index to ensure iteration
        df_rows = df_sql.copy()

        # helper to evaluate dfcolumn expressions like '[DST-1]/0.9'
        expr_re = re.compile(r"\[([^\]]+)\](?:\s*([*/+-])\s*([0-9.]+))?")

        def eval_expr(expr: str, row: pd.Series):
            # direct column name (case-insensitive)
            if expr in row.index:
                return row.get(expr)
            if expr.lower() in row.index:
                return row.get(expr.lower())

            # expression with bracketed category
            m = expr_re.search(expr)
            if not m:
                return None
            cat = m.group(1)
            op = m.group(2)
            num = m.group(3)
            # try several casings for the category column
            val = None
            if cat in row.index:
                val = row.get(cat)
            elif cat.lower() in row.index:
                val = row.get(cat.lower())
            elif cat.upper() in row.index:
                val = row.get(cat.upper())
            try:
                if pd.isna(val):
                    return None
            except Exception:
                pass
            try:
                if op and num:
                    numf = float(num)
                    if op == "/":
                        return float(val) / numf
                    if op == "*":
                        return float(val) * numf
                    if op == "+":
                        return float(val) + numf
                    if op == "-":
                        return float(val) - numf
                return val
            except Exception:
                return val

        # start writing data rows beginning at starting_row
        write_row = starting_row
        for _, prow in df_rows.iterrows():
            for col in schema_columns:
                cname = col.get("columnname")
                dfield = col.get("dfcolumnname", "")
                if not dfield:
                    continue

                # map common direct names
                val = None
                dfield_l = dfield.strip()
                if dfield_l in ("partid", "sku", "description", "effective_date"):
                    val = prow.get(dfield_l)
                else:
                    # expression or pivoted category like [RTL-1]
                    val = eval_expr(dfield_l, prow)

                col_idx = col_map.get(cname)
                if col_idx is None:
                    # append to the end
                    col_idx = ws.max_column + 1
                    col_map[cname] = col_idx

                ws.cell(row=write_row, column=col_idx, value=val)

            write_row += 1

        wb.save(dest)
        typer.echo(f"Processed and wrote file: {dest}")


# ----------------------------------------------------------------
# Main
# ----------------------------------------------------------------
if __name__ == "__main__":
    app()
